#!/usr/bin/env python3
"""Add a `question_number` column to runs/final-combined.csv using the global `#`
from Gloden Evaluation Set1.xlsx. Matches on (user_email, question text).
"""
import csv
import os
import re
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "Gloden Evaluation Set1.xlsx")
IN_CSV = os.path.join(ROOT, "runs/final-combined.csv")
OUT_CSV = os.path.join(ROOT, "runs/final-combined.csv")


def norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def load_xlsx_index():
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    by_email_q = {}
    by_q = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, email, question = row[0], row[1], row[2]
        if question is None:
            continue
        try:
            num = int(num)
        except (TypeError, ValueError):
            continue
        ek = norm(email)
        qk = norm(question)
        by_email_q[(ek, qk)] = num
        by_q.setdefault(qk, num)
    return by_email_q, by_q


def main():
    by_email_q, by_q = load_xlsx_index()

    with open(IN_CSV) as fp:
        reader = csv.reader(fp)
        header = next(reader)
        all_rows = list(reader)

    if "question_number" in header:
        print("question_number column already exists; nothing to do.")
        return

    # Insert question_number right after row_number (or at the front if absent).
    try:
        insert_at = header.index("row_number") + 1
    except ValueError:
        insert_at = 0
    new_header = header[:insert_at] + ["question_number"] + header[insert_at:]

    q_idx = header.index("question")
    email_idx = header.index("user_email")
    user_idx = header.index("user") if "user" in header else None

    out_rows = []
    matched = 0
    unmatched = 0
    for row in all_rows:
        # Preserve the pseudo-subheader row where values echo the agent group label.
        if user_idx is not None and row[user_idx] == "user":
            out_rows.append(row[:insert_at] + ["question_number"] + row[insert_at:])
            continue

        email = norm(row[email_idx])
        question = norm(row[q_idx])
        qnum = by_email_q.get((email, question)) or by_q.get(question)

        if qnum is None:
            unmatched += 1
            value = ""
        else:
            matched += 1
            value = str(qnum)

        out_rows.append(row[:insert_at] + [value] + row[insert_at:])

    with open(OUT_CSV, "w", newline="") as fp:
        w = csv.writer(fp)
        w.writerow(new_header)
        w.writerows(out_rows)

    print(f"wrote {OUT_CSV}")
    print(f"  matched={matched} unmatched={unmatched}")
    print(f"  column inserted at position {insert_at} (0-indexed)")


if __name__ == "__main__":
    main()
