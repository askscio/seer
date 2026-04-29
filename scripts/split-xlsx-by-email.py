#!/usr/bin/env python3
"""
Split Gloden Evaluation Set1.xlsx into one file per user email.

Each output file has the same headers as the original, preserves original
row numbers in the '#' column, and contains only that user's rows.

Naming uses the first-name slug, matching 'Gloden Evaluation Set1 - niti.xlsx'
format so the existing orchestrator picks it up without changes.

Usage:
    python3 scripts/split-xlsx-by-email.py \
        --input "Gloden Evaluation Set1.xlsx" \
        --out-dir .
"""

import argparse
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook, Workbook


def slug_for(email: str) -> str:
    """Map 'niti.jain@ericsson.com' -> 'niti'."""
    local = email.split("@", 1)[0].strip()
    first = re.split(r"[.\-_+]", local)[0]
    return first.lower()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="Source XLSX path")
    parser.add_argument("--out-dir", default=".", help="Where to write per-user XLSX files")
    parser.add_argument(
        "--base-name",
        default="Gloden Evaluation Set1",
        help="Output file base name; actual name becomes '<base> - <slug>.xlsx'",
    )
    args = parser.parse_args()

    src = Path(args.input)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("Empty input workbook")

    headers = [h for h in rows[0]]
    normalized = [str(h).strip().lower() if h is not None else "" for h in headers]
    try:
        email_idx = normalized.index("user email id")
    except ValueError:
        email_idx = next(i for i, h in enumerate(normalized) if "email" in h)

    # Group rows by email (stripped, lower-cased).
    buckets: dict[str, list[tuple]] = defaultdict(list)
    for row in rows[1:]:
        if row is None:
            continue
        raw_email = row[email_idx]
        if raw_email is None or str(raw_email).strip() == "":
            continue
        email = str(raw_email).strip()
        buckets[email].append(row)

    print(f"Found {sum(len(v) for v in buckets.values())} rows across {len(buckets)} emails.")

    # Avoid slug collisions between emails.
    used_slugs: dict[str, str] = {}
    for email in sorted(buckets.keys()):
        base_slug = slug_for(email)
        slug = base_slug
        suffix = 2
        while slug in used_slugs and used_slugs[slug] != email:
            slug = f"{base_slug}{suffix}"
            suffix += 1
        used_slugs[slug] = email

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.append(list(headers))
        for row in buckets[email]:
            out_ws.append(list(row))

        out_path = out_dir / f"{args.base_name} - {slug}.xlsx"
        out_wb.save(out_path)
        print(f"  {slug:10s} -> {len(buckets[email]):3d} rows  -> {out_path}")


if __name__ == "__main__":
    main()
